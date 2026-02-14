import os
import tempfile
import unittest

import db
import export
from helpers import calc_days, calc_interest_rate_pa


class ExportTests(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmpdir.cleanup)

        # Redirect DB to temp directory
        db_path = os.path.join(self.tmpdir.name, "test_export.db")
        self._orig_db_path = db.DB_PATH
        db.DB_PATH = db_path
        self.addCleanup(setattr, db, "DB_PATH", self._orig_db_path)
        db.init_db()

        # Redirect export path to temp directory
        export_dir = os.path.join(self.tmpdir.name, "export")
        self._orig_export_path = export.EXPORT_PATH
        self._orig_export_file = export.EXPORT_FILE
        export.EXPORT_PATH = export_dir
        export.EXPORT_FILE = os.path.join(export_dir, "tenordash.xlsx")
        self.addCleanup(setattr, export, "EXPORT_PATH", self._orig_export_path)
        self.addCleanup(setattr, export, "EXPORT_FILE", self._orig_export_file)

        self.export_file = export.EXPORT_FILE

    def _seed_data(self):
        """Insert a bank, credit line, and advance for testing."""
        conn = db.get_db()
        try:
            conn.execute(
                "INSERT INTO banks (bank_key, bank_name) VALUES (?, ?)",
                ("B001", "Test Bank"),
            )
            conn.execute(
                "INSERT INTO credit_lines (id, bank_key, description, currency, amount, "
                "committed, start_date, end_date, note, archived) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                ("CL001", "B001", "Test Facility", "CHF", 100_000_000,
                 "Yes", "2026-01-01", "2027-01-01", "test note", 0),
            )
            conn.execute(
                "INSERT INTO fixed_advances (id, bank, credit_line_id, start_date, end_date, "
                "continuation_date, currency, amount_original, interest_amount) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                ("FV0001", "B001", "CL001", "2026-01-15", "2026-04-15",
                 "2026-04-10", "CHF", 50_000_000, 125_000.0),
            )
            conn.commit()
        finally:
            conn.close()

    def test_creates_file(self):
        self._seed_data()
        export.export_xlsx()
        self.assertTrue(os.path.isfile(self.export_file))

    def test_creates_export_directory(self):
        self._seed_data()
        export.export_xlsx()
        self.assertTrue(os.path.isdir(export.EXPORT_PATH))

    def test_sheet_names(self):
        from openpyxl import load_workbook
        self._seed_data()
        export.export_xlsx()
        wb = load_workbook(self.export_file)
        self.assertEqual(wb.sheetnames, ["tblFV", "tblCreditLines"])
        wb.close()

    def test_advance_columns(self):
        from openpyxl import load_workbook
        self._seed_data()
        export.export_xlsx()
        wb = load_workbook(self.export_file)
        ws = wb["tblFV"]
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers, export.ADVANCE_COLUMNS)
        wb.close()

    def test_credit_line_columns(self):
        from openpyxl import load_workbook
        self._seed_data()
        export.export_xlsx()
        wb = load_workbook(self.export_file)
        ws = wb["tblCreditLines"]
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers, export.CREDIT_LINE_COLUMNS)
        wb.close()

    def test_row_counts(self):
        from openpyxl import load_workbook
        self._seed_data()
        export.export_xlsx()
        wb = load_workbook(self.export_file)
        # 1 header + 1 data row
        self.assertEqual(wb["tblFV"].max_row, 2)
        self.assertEqual(wb["tblCreditLines"].max_row, 2)
        wb.close()

    def test_calculated_fields(self):
        from openpyxl import load_workbook
        self._seed_data()
        export.export_xlsx()
        wb = load_workbook(self.export_file)
        ws = wb["tblFV"]
        # Row 2 is the data row
        row = [cell.value for cell in ws[2]]
        headers = [cell.value for cell in ws[1]]
        data = dict(zip(headers, row))

        expected_days = calc_days("2026-01-15", "2026-04-15")
        expected_rate = calc_interest_rate_pa(125_000.0, 50_000_000, expected_days)

        self.assertEqual(data["days"], expected_days)
        self.assertAlmostEqual(data["rate_pa"], round(expected_rate, 6), places=6)
        wb.close()

    def test_empty_tables(self):
        from openpyxl import load_workbook
        # No seed data — tables are empty
        export.export_xlsx()
        self.assertTrue(os.path.isfile(self.export_file))
        wb = load_workbook(self.export_file)
        # Only header rows
        self.assertEqual(wb["tblFV"].max_row, 1)
        self.assertEqual(wb["tblCreditLines"].max_row, 1)
        wb.close()

    def test_includes_archived_credit_lines(self):
        from openpyxl import load_workbook
        self._seed_data()
        # Add an archived credit line
        conn = db.get_db()
        try:
            conn.execute(
                "INSERT INTO credit_lines (id, bank_key, description, currency, amount, "
                "committed, start_date, end_date, note, archived) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                ("CL002", "B001", "Archived Facility", "EUR", 50_000_000,
                 "No", "2025-01-01", "2025-12-31", None, 1),
            )
            conn.commit()
        finally:
            conn.close()

        export.export_xlsx()
        wb = load_workbook(self.export_file)
        # Header + 2 credit lines (active + archived)
        self.assertEqual(wb["tblCreditLines"].max_row, 3)
        wb.close()


if __name__ == "__main__":
    unittest.main()
