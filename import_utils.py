from datetime import date, datetime

from openpyxl import load_workbook


# Column mappings: Excel header → DB field
CL_COLUMN_MAP = {
    "Credit Line ID": "id",
    "BankKey": "bank_key",
    "Description": "description",
    "Currency": "currency",
    "Amount": "amount",
    "Committed": "committed",
    "Start Date": "start_date",
    "End Date": "end_date",
    "Note": "note",
}

ADV_COLUMN_MAP = {
    "ID": "id",
    "Bank": "bank",
    "Linked\nCredit Line": "credit_line_id",
    "Linked Credit Line": "credit_line_id",
    "Start Date": "start_date",
    "End Date": "end_date",
    "Continuation Date": "continuation_date",
    "Currency": "currency",
    "Amount Original": "amount_original",
    "Interest Amount": "interest_amount",
}

CL_REQUIRED = {"id", "bank_key", "currency", "amount", "committed", "start_date"}
ADV_REQUIRED = {"id", "bank", "credit_line_id", "start_date", "end_date",
                "continuation_date", "currency", "amount_original", "interest_amount"}


def _normalize_date(value):
    """Convert Excel date or string to ISO format string."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    s = str(value).strip()
    if not s or s == "--":
        return None
    # Try ISO parse
    try:
        return date.fromisoformat(s).isoformat()
    except ValueError:
        return s  # leave as-is, will be caught by validation


def _normalize_amount(value):
    """Convert to numeric, handling None and string formats."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    s = str(value).replace(",", "").strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _parse_sheet(ws, column_map, header_row=4):
    """Parse a worksheet using column_map to rename headers. Returns list of dicts."""
    headers_raw = [cell.value for cell in ws[header_row]]
    # Map Excel headers to DB field names
    col_indices = {}
    for idx, header in enumerate(headers_raw):
        if header and header.strip() in column_map:
            col_indices[idx] = column_map[header.strip()]
        elif header and header in column_map:
            col_indices[idx] = column_map[header]

    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        # Skip completely empty rows
        if all(v is None for v in row):
            continue
        record = {}
        for idx, field in col_indices.items():
            record[field] = row[idx] if idx < len(row) else None
        # Skip rows where ID is empty (padding rows)
        id_field = "id" if "id" in record else None
        if id_field and not record.get(id_field):
            continue
        rows.append(record)

    return rows


def _normalize_credit_line(row):
    """Normalize a credit line row's types."""
    row["start_date"] = _normalize_date(row.get("start_date"))
    row["end_date"] = _normalize_date(row.get("end_date"))
    row["amount"] = _normalize_amount(row.get("amount"))
    if row["amount"] is not None:
        row["amount"] = int(row["amount"])
    # Normalize committed to Yes/No
    committed = str(row.get("committed", "")).strip()
    row["committed"] = "Yes" if committed.lower() in ("yes", "y", "1", "true") else "No"
    # Strip string fields
    for f in ("id", "bank_key", "description", "currency", "note"):
        if row.get(f) is not None:
            row[f] = str(row[f]).strip()
    return row


def _normalize_advance(row):
    """Normalize an advance row's types."""
    for f in ("start_date", "end_date", "continuation_date"):
        row[f] = _normalize_date(row.get(f))
    row["amount_original"] = _normalize_amount(row.get("amount_original"))
    if row["amount_original"] is not None:
        row["amount_original"] = int(row["amount_original"])
    row["interest_amount"] = _normalize_amount(row.get("interest_amount"))
    if row["interest_amount"] is not None:
        row["interest_amount"] = float(row["interest_amount"])
    # Strip string fields
    for f in ("id", "bank", "credit_line_id", "currency"):
        if row.get(f) is not None:
            row[f] = str(row[f]).strip()
    return row


def _extract_banks(credit_lines, advances):
    """Extract unique banks from credit lines (bank_key) and advances (bank name)."""
    bank_keys = {}
    for cl in credit_lines:
        bk = cl.get("bank_key")
        if bk:
            bank_keys.setdefault(bk, None)

    # Build credit_line_id → bank_key lookup
    cl_to_bk = {cl["id"]: cl["bank_key"] for cl in credit_lines if cl.get("id") and cl.get("bank_key")}

    # From advances, map bank names to bank_keys via credit_line_id
    for adv in advances:
        cl_id = adv.get("credit_line_id")
        bank_name = adv.get("bank")
        if cl_id and bank_name and cl_id in cl_to_bk:
            bk = cl_to_bk[cl_id]
            if bk in bank_keys and bank_keys[bk] is None:
                bank_keys[bk] = bank_name

    # Fill any remaining None names with bank_key as fallback
    rows = []
    for bk, name in bank_keys.items():
        rows.append({"bank_key": bk, "bank_name": name or bk})

    return rows


def validate_credit_line(row):
    """Validate a credit line row. Returns list of error strings."""
    errors = []
    for field in CL_REQUIRED:
        if not row.get(field):
            errors.append(f"Missing required field: {field}")
    if row.get("amount") is not None and not isinstance(row["amount"], (int, float)):
        errors.append("Amount must be numeric")
    if row.get("start_date"):
        try:
            date.fromisoformat(row["start_date"])
        except (ValueError, TypeError):
            errors.append(f"Invalid start_date: {row['start_date']}")
    return errors


def validate_advance(row):
    """Validate an advance row. Returns list of error strings."""
    errors = []
    for field in ADV_REQUIRED:
        if not row.get(field) and row.get(field) != 0:
            errors.append(f"Missing required field: {field}")
    if row.get("amount_original") is not None and not isinstance(row["amount_original"], (int, float)):
        errors.append("Amount must be numeric")
    if row.get("interest_amount") is not None and not isinstance(row["interest_amount"], (int, float)):
        errors.append("Interest amount must be numeric")
    # Date order check
    try:
        s = date.fromisoformat(row.get("start_date", ""))
        e = date.fromisoformat(row.get("end_date", ""))
        if e <= s:
            errors.append("end_date must be later than start_date")
    except (ValueError, TypeError):
        pass  # Already caught by required field check
    return errors


def parse_excel(filepath):
    """Parse an Excel file and return structured data for preview.

    Returns dict with keys: banks, credit_lines, advances.
    Each value is {"rows": [...], "errors": [{"row": N, "messages": [...]}]}.
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    try:
        result = {
            "banks": {"rows": [], "errors": []},
            "credit_lines": {"rows": [], "errors": []},
            "advances": {"rows": [], "errors": []},
        }

        # Parse Credit Lines sheet
        cl_sheet_name = None
        for name in wb.sheetnames:
            if "credit" in name.lower() and "line" in name.lower():
                cl_sheet_name = name
                break
        if cl_sheet_name:
            raw_cls = _parse_sheet(wb[cl_sheet_name], CL_COLUMN_MAP)
            for i, row in enumerate(raw_cls):
                row = _normalize_credit_line(row)
                errors = validate_credit_line(row)
                if errors:
                    result["credit_lines"]["errors"].append({"row": i + 5, "messages": errors})
                else:
                    result["credit_lines"]["rows"].append(row)

        # Parse Fixed Advances sheet
        adv_sheet_name = None
        for name in wb.sheetnames:
            if "advance" in name.lower() or "fixed" in name.lower():
                adv_sheet_name = name
                break
        if adv_sheet_name:
            raw_advs = _parse_sheet(wb[adv_sheet_name], ADV_COLUMN_MAP)
            for i, row in enumerate(raw_advs):
                row = _normalize_advance(row)
                errors = validate_advance(row)
                if errors:
                    result["advances"]["errors"].append({"row": i + 5, "messages": errors})
                else:
                    result["advances"]["rows"].append(row)

        # Extract banks from both sheets
        result["banks"]["rows"] = _extract_banks(
            result["credit_lines"]["rows"],
            result["advances"]["rows"],
        )

        return result
    finally:
        wb.close()
